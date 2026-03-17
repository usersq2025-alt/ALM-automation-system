import streamlit as st
import pandas as pd
import io
import zipfile
import xlsxwriter

st.set_page_config(
    page_title="أداة مقرأة",
    page_icon="📖",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;900&display=swap');
    html, body, [class*="css"] { font-family: 'Tajawal', sans-serif; direction: rtl; }
    .stApp { background: linear-gradient(135deg, #f8f4ef 0%, #ede8e0 100%); }
    h1, h2, h3 { font-family: 'Tajawal', sans-serif !important; }
    .hero-header {
        background: linear-gradient(135deg, #2d5016 0%, #4a7c28 50%, #6ba535 100%);
        border-radius: 16px; padding: 2rem 2.5rem; margin-bottom: 2rem;
        box-shadow: 0 8px 32px rgba(45,80,22,0.25); text-align: center; color: white;
    }
    .hero-header h1 {
        font-size: 2.4rem; font-weight: 900; margin: 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.2);
    }
    .hero-header p { font-size: 1.05rem; margin: 0.5rem 0 0; opacity: 0.88; font-weight: 300; }
    .stat-card {
        background: white; border-radius: 12px; padding: 1.2rem 1.5rem;
        box-shadow: 0 2px 12px rgba(0,0,0,0.07); border-right: 4px solid #4a7c28; margin-bottom: 1rem;
    }
    .stat-card .number { font-size: 2rem; font-weight: 900; color: #2d5016; line-height: 1; }
    .stat-card .label { font-size: 0.85rem; color: #777; margin-top: 4px; }
    .file-chip {
        display: inline-block; background: #e8f5e0; border: 1px solid #a8d878;
        color: #2d5016; border-radius: 20px; padding: 4px 14px;
        font-size: 0.82rem; margin: 3px; font-weight: 500;
    }
    .success-banner {
        background: linear-gradient(90deg, #e8f5e0, #d4edbe); border: 1px solid #a8d878;
        border-radius: 10px; padding: 1rem 1.5rem; color: #2d5016;
        font-weight: 600; font-size: 1.05rem; margin: 1rem 0;
    }
    .section-title {
        font-size: 1.1rem; font-weight: 700; color: #2d5016;
        border-bottom: 2px solid #a8d878; padding-bottom: 6px; margin: 1.5rem 0 1rem;
    }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a3a08 0%, #2d5016 100%) !important;
    }
    [data-testid="stSidebar"] * { color: #d8f0b8 !important; }
    [data-testid="stSidebar"] .stTextArea textarea {
        background: rgba(255,255,255,0.1) !important;
        border: 1px solid rgba(168,216,120,0.4) !important;
        color: #f0f8e8 !important;
        font-family: 'Tajawal', sans-serif !important;
        direction: rtl;
    }
    [data-testid="stSidebar"] label { font-weight: 600 !important; font-size: 0.9rem !important; }
    .stButton > button {
        font-family: 'Tajawal', sans-serif !important;
        font-weight: 700 !important; border-radius: 10px !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #2d5016, #4a7c28) !important;
        border: none !important; box-shadow: 0 4px 15px rgba(45,80,22,0.3) !important;
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #1a5276, #2874a6) !important;
        color: white !important; font-family: 'Tajawal', sans-serif !important;
        font-weight: 700 !important; border: none !important;
        border-radius: 10px !important; padding: 0.6rem 2rem !important;
        font-size: 1rem !important; box-shadow: 0 4px 15px rgba(26,82,118,0.3) !important;
    }
    .upload-zone {
        background: white; border: 2px dashed #a8d878; border-radius: 16px;
        padding: 2rem; text-align: center; margin: 1rem 0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def get_short_name(full_name):
    parts = str(full_name).strip().split()
    if len(parts) >= 2:
        return parts[0] + "." + parts[1][0]
    return parts[0] if parts else full_name


def parse_list(text):
    return [line.strip() for line in text.strip().splitlines() if line.strip()]


def build_excel(df, days, periods, statuses):
    output = io.BytesIO()

    columns_order = [
        "الرقم", "الاسم", "رقم الواتس اب", "المجموعة",
        "البلد", "المواليد", "الإجازة", "المعلمة",
        "الحالة", "يوم الاختبار", "توقيت الاختبار", "الفترة", "الملاحظات",
    ]

    for col in columns_order:
        if col not in df.columns:
            df[col] = ""

    df = df[columns_order]
    num_rows = len(df)
    extra_rows = 50

    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = workbook.add_worksheet("الطالبات")
    ws.right_to_left()

    header_fmt = workbook.add_format({
        "bold": True, "font_name": "Tajawal", "font_size": 11,
        "align": "center", "valign": "vcenter",
        "fg_color": "#2d5016", "font_color": "white",
        "border": 1, "text_wrap": True, "locked": True,
    })
    locked_fmt = workbook.add_format({
        "font_name": "Tajawal", "font_size": 10,
        "align": "center", "valign": "vcenter",
        "border": 1, "locked": True, "bg_color": "#f5f5f5",
    })
    unlocked_fmt = workbook.add_format({
        "font_name": "Tajawal", "font_size": 10,
        "align": "center", "valign": "vcenter",
        "border": 1, "locked": False, "bg_color": "#fffde7",
    })
    unlocked_extra_fmt = workbook.add_format({
        "font_name": "Tajawal", "font_size": 10,
        "locked": False, "bg_color": "#fff9e6",
    })

    col_widths = [8, 28, 18, 16, 12, 12, 12, 20, 22, 18, 18, 18, 25]
    for i, w in enumerate(col_widths):
        ws.set_column(i, i, w)

    ws.set_row(0, 30)
    for col_idx, col_name in enumerate(columns_order):
        ws.write(0, col_idx, col_name, header_fmt)

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 1
        ws.set_row(excel_row, 20)
        for col_idx, col_name in enumerate(columns_order):
            val = row[col_name]
            val = "" if pd.isna(val) else val
            fmt = locked_fmt if col_idx < 8 else unlocked_fmt
            ws.write(excel_row, col_idx, val, fmt)

    for extra in range(extra_rows):
        excel_row = num_rows + 1 + extra
        ws.set_row(excel_row, 20)
        for col_idx in range(13):
            fmt = locked_fmt if col_idx < 8 else unlocked_extra_fmt
            ws.write(excel_row, col_idx, "", fmt)

    last_val_row = num_rows + extra_rows

    ws.data_validation(1, 8, last_val_row, 8, {
        "validate": "list", "source": statuses,
        "input_title": "الحالة", "input_message": "اختاري الحالة المناسبة",
    })
    ws.data_validation(1, 9, last_val_row, 9, {
        "validate": "list", "source": days,
        "input_title": "يوم الاختبار", "input_message": "اختاري اليوم",
    })
    ws.data_validation(1, 11, last_val_row, 11, {
        "validate": "list", "source": periods,
        "input_title": "الفترة", "input_message": "اختاري الفترة",
    })

    ws.protect("", {
        "sheet": True, "insert_rows": True, "insert_columns": False,
        "delete_rows": False, "sort": False, "autofilter": False,
        "select_locked_cells": True, "select_unlocked_cells": True,
    })

    workbook.close()
    output.seek(0)
    return output.read()


def process_files(uploaded_files, days, periods, statuses):
    results = {}
    errors = []

    for uf in uploaded_files:
        try:
            file_bytes = uf.read()

            if uf.name.lower().endswith(".csv"):
                df = pd.read_csv(io.BytesIO(file_bytes))
            elif uf.name.lower().endswith(".xls"):
                df = pd.read_excel(io.BytesIO(file_bytes), engine="xlrd")
            else:
                # .xlsx — read data only, ignore all styles to avoid Fill errors
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows:
                    raise ValueError("الملف فارغ")
                headers = [str(c) if c is not None else "" for c in rows[0]]
                data = [list(r) for r in rows[1:]]
                df = pd.DataFrame(data, columns=headers)

            teacher_col = next((c for c in df.columns if "المعلمة" in str(c)), None)
            if teacher_col and not df[teacher_col].dropna().empty:
                raw_name = str(df[teacher_col].dropna().iloc[0]).strip()
                short = get_short_name(raw_name)
            else:
                short = uf.name.rsplit(".", 1)[0]

            xlsx_bytes = build_excel(df.copy(), days, periods, statuses)
            out_name = short + ".xlsx"
            base = out_name
            counter = 1
            while out_name in results:
                out_name = base.replace(".xlsx", "_" + str(counter) + ".xlsx")
                counter += 1

            results[out_name] = xlsx_bytes

        except Exception as e:
            errors.append("❌ " + uf.name + ": " + str(e))

    return results, errors


# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        """
        <div style='text-align:center; padding:1rem 0 0.5rem;'>
            <div style='font-size:2.5rem'>📖</div>
            <div style='font-size:1.2rem; font-weight:900; color:#d8f0b8;'>إعدادات الدورة</div>
            <div style='font-size:0.8rem; color:#a8d878; margin-top:4px;'>خصّصي القيم لكل دورة</div>
        </div>
        <hr style='border-color:rgba(168,216,120,0.3); margin:0.8rem 0;'>
        """,
        unsafe_allow_html=True,
    )

    days_text = st.text_area(
        "📅 أيام الأسبوع",
        value="الاثنين 3/23\nالثلاثاء 3/24\nالأربعاء 3/25\nالخميس 3/26\nالجمعة 3/27\nالسبت 3/28\nالأحد 3/29",
        height=160,
        help="كل يوم في سطر منفصل",
    )
    periods_text = st.text_area(
        "⏰ الفترات",
        value="فجراً من 5.45-9.00\nضحى 9:15-12.30\nظهراً 12:45-4.15\nعصراً 4.30-7.00\nليلاً 7.15-9.30",
        height=140,
        help="كل فترة في سطر منفصل",
    )
    statuses_text = st.text_area(
        "📋 قائمة الحالات",
        value="أنهت المقرر\nلم تنه المقرر\nساكنة\nمنسحبة\nأخرجتها الإدارة لأنها مخالفة\nلا يوجد واتس\nتم نقلها لغير مجموعة",
        height=175,
        help="كل حالة في سطر منفصل",
    )

    days_list = parse_list(days_text)
    periods_list = parse_list(periods_text)
    statuses_list = parse_list(statuses_text)

    st.markdown(
        "<div style='margin-top:1rem; padding:0.8rem; background:rgba(255,255,255,0.08);"
        "border-radius:8px; font-size:0.82rem; color:#a8d878;'>"
        "✅ " + str(len(days_list)) + " أيام &nbsp;|&nbsp; ✅ "
        + str(len(periods_list)) + " فترات &nbsp;|&nbsp; ✅ "
        + str(len(statuses_list)) + " حالة</div>",
        unsafe_allow_html=True,
    )


# ── Main ───────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="hero-header">
        <h1>📖 أداة أتمتة جداول مقرأة</h1>
        <p>ارفعي ملفات Excel أو CSV الخام وستحصلين على جداول منسقة، محمية، وجاهزة للمعلمات</p>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="section-title">📂 رفع الملفات</div>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "اسحبي الملفات هنا أو اضغطي للاختيار",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if uploaded_files:
    cols = st.columns(4)
    with cols[0]:
        st.markdown(
            '<div class="stat-card"><div class="number">' + str(len(uploaded_files))
            + '</div><div class="label">ملف مرفوع</div></div>',
            unsafe_allow_html=True,
        )
    with cols[1]:
        st.markdown(
            '<div class="stat-card"><div class="number">' + str(len(days_list))
            + '</div><div class="label">أيام الاختبار</div></div>',
            unsafe_allow_html=True,
        )
    with cols[2]:
        st.markdown(
            '<div class="stat-card"><div class="number">' + str(len(periods_list))
            + '</div><div class="label">فترة متاحة</div></div>',
            unsafe_allow_html=True,
        )
    with cols[3]:
        st.markdown(
            '<div class="stat-card"><div class="number">' + str(len(statuses_list))
            + '</div><div class="label">حالة في القائمة</div></div>',
            unsafe_allow_html=True,
        )

    chips = " ".join(
        ['<span class="file-chip">📄 ' + f.name + "</span>" for f in uploaded_files]
    )
    st.markdown(
        "<div style='margin:0.5rem 0 1.5rem'>" + chips + "</div>",
        unsafe_allow_html=True,
    )

    if st.button("⚡ معالجة الملفات وتوليد جداول المعلمات", type="primary", use_container_width=True):
        with st.spinner("جارٍ المعالجة..."):
            results, errors = process_files(uploaded_files, days_list, periods_list, statuses_list)

        for e in errors:
            st.error(e)

        if results:
            st.markdown(
                '<div class="success-banner">✅ تمت معالجة ' + str(len(results)) + ' ملف بنجاح!</div>',
                unsafe_allow_html=True,
            )

            st.markdown('<div class="section-title">👁️ معاينة الملفات الناتجة</div>', unsafe_allow_html=True)
            preview_data = [{"اسم الملف الناتج": fname, "الحالة": "✅ جاهز"} for fname in results]
            st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname, fbytes in results.items():
                    zf.writestr(fname, fbytes)
            zip_buffer.seek(0)

            st.markdown('<div class="section-title">📦 تحميل الملفات</div>', unsafe_allow_html=True)
            st.download_button(
                label="⬇️ تحميل جميع الملفات (" + str(len(results)) + " ملف) — ZIP",
                data=zip_buffer,
                file_name="جداول_المعلمات.zip",
                mime="application/zip",
                use_container_width=True,
            )

else:
    st.markdown(
        """
        <div class="upload-zone">
            <div style="font-size:3rem; margin-bottom:0.5rem">📂</div>
            <div style="font-size:1.1rem; font-weight:600; color:#4a7c28;">لم يتم رفع أي ملفات بعد</div>
            <div style="font-size:0.85rem; color:#888; margin-top:0.3rem;">يدعم الصيغ: xlsx, xls, csv</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <hr style="margin:2rem 0 1rem; border-color:#d0e8c0;">
    <div style="text-align:center; color:#999; font-size:0.8rem; font-family:'Tajawal',sans-serif;">
        أداة مقرأة — مبنية بـ Python & Streamlit &nbsp;|&nbsp; 📖
    </div>
    """,
    unsafe_allow_html=True,
)
